"""
Shared source-document to page-image preprocessing helpers.
"""

from __future__ import annotations

import hashlib
import importlib.util
import os
import re
import shutil
import textwrap
from pathlib import Path
from typing import Any, Iterable, Sequence

DEFAULT_PAGE_BUNDLE_VERSION = "1.0"
IMAGE_SOURCE_SUFFIXES = frozenset({".png", ".jpg", ".jpeg", ".webp", ".gif"})
DOCUMENT_SOURCE_SUFFIXES = frozenset({".pdf", ".docx", ".xlsx"})
RENDERABLE_SOURCE_SUFFIXES = IMAGE_SOURCE_SUFFIXES | DOCUMENT_SOURCE_SUFFIXES


class MissingDependencyError(RuntimeError):
    """Raised when an optional rendering dependency is unavailable."""


def enumerate_renderable_documents(
    root: Path | str,
    *,
    exclude_paths: Sequence[Path | str] | None = None,
    recursive: bool = True,
) -> dict[str, Any]:
    """
    Discover renderable source files under ``root``.

    JSON and other unsupported files are reported in ``skipped`` rather than
    silently ignored so callers can audit corpus coverage.

    Returns:
        Dict with keys:

        - ``root`` (str): Resolved source root path.
        - ``documents`` (List[Path]): Renderable source files found.
        - ``skipped`` (List[dict]): Each entry has ``path`` and ``reason``
          (e.g. ``"excluded"``, ``"unsupported_suffix:.json"``).
    """

    source_root = Path(root).expanduser().resolve()
    excluded = {
        Path(path).expanduser().resolve()
        for path in (exclude_paths or [])
    }
    discovered: list[Path] = []
    skipped: list[dict[str, str]] = []

    if not source_root.exists():
        return {
            "root": str(source_root),
            "documents": discovered,
            "skipped": [{"path": str(source_root), "reason": "missing_path"}],
        }

    candidates: Iterable[Path]
    if source_root.is_file():
        candidates = [source_root]
    elif recursive:
        candidates = sorted(path.resolve() for path in source_root.rglob("*") if path.is_file())
    else:
        candidates = sorted(path.resolve() for path in source_root.iterdir() if path.is_file())

    for path in candidates:
        if path in excluded:
            skipped.append({"path": str(path), "reason": "excluded"})
            continue
        suffix = path.suffix.lower()
        if suffix in RENDERABLE_SOURCE_SUFFIXES:
            discovered.append(path)
        else:
            skipped.append({"path": str(path), "reason": f"unsupported_suffix:{suffix or 'none'}"})

    return {
        "root": str(source_root),
        "documents": discovered,
        "skipped": skipped,
    }


def render_documents(
    documents: Sequence[Path | str],
    output_dir: Path | str,
    *,
    source_root: Path | str | None = None,
) -> dict[str, Any]:
    """
    Render supported documents into page-image assets and PageBundle-like JSON.

    Returns:
        Dict with keys:

        - ``status`` (str): ``"passed"`` if any pages rendered, ``"skipped"`` otherwise.
        - ``output_dir`` (str): Resolved output directory.
        - ``bundles`` (List[dict]): Per-document bundles with ``doc_id``,
          ``source_uri``, ``metadata``, and ``pages`` list.
        - ``rendered`` (List[dict]): Flat list of all page dicts. Each has
          ``source``, ``image_path``, ``page_number``, ``page_id``,
          ``doc_id``, ``renderer``, ``text``, ``width``, ``height``.
        - ``skipped`` (List[dict]): Files that couldn't be rendered
          (``path``, ``reason``).
        - ``summary`` (dict): Aggregate counts: ``documents_requested``,
          ``documents_rendered``, ``pages_rendered``, ``by_source``,
          ``renderer_counts``.
    """

    output_root = Path(output_dir).expanduser().resolve()
    output_root.mkdir(parents=True, exist_ok=True)
    resolved_documents = [Path(document).expanduser().resolve() for document in documents]
    effective_source_root = _resolve_source_root(resolved_documents, source_root)

    bundles: list[dict[str, Any]] = []
    rendered: list[dict[str, Any]] = []
    skipped: list[dict[str, str]] = []
    by_source: list[dict[str, Any]] = []
    renderer_counts: dict[str, int] = {}

    for source in resolved_documents:
        if not source.exists():
            skipped.append({"path": str(source), "reason": "missing_path"})
            continue
        if not source.is_file():
            skipped.append({"path": str(source), "reason": "not_a_file"})
            continue

        source.suffix.lower()
        doc_id = _stable_doc_id(source, effective_source_root)
        document_dir = output_root / doc_id
        if document_dir.exists():
            shutil.rmtree(document_dir)
        document_dir.mkdir(parents=True, exist_ok=True)
        try:
            pages = _render_source_pages(source, document_dir)
        except MissingDependencyError as exc:
            skipped.append({"path": str(source), "reason": f"missing_dependency:{exc}"})
            continue
        except Exception as exc:  # pragma: no cover - filesystem and optional deps
            skipped.append({"path": str(source), "reason": f"render_failed:{exc}"})
            continue

        relative_source = _relative_source_name(source, effective_source_root)
        bundle = _build_page_bundle(
            doc_id=doc_id,
            source=source,
            relative_source=relative_source,
            pages=pages,
        )
        bundles.append(bundle)
        flat_pages = _flatten_bundle_pages(bundle)
        rendered.extend(flat_pages)
        by_source.append(
            {
                "doc_id": doc_id,
                "source_path": str(source),
                "page_count": len(flat_pages),
                "renderers": sorted({page["renderer"] for page in flat_pages}),
            }
        )
        for page in flat_pages:
            renderer_counts[page["renderer"]] = renderer_counts.get(page["renderer"], 0) + 1

    return {
        "status": "passed" if rendered else "skipped",
        "output_dir": str(output_root),
        "bundles": bundles,
        "rendered": rendered,
        "skipped": skipped,
        "summary": {
            "documents_requested": len(resolved_documents),
            "documents_rendered": len(bundles),
            "pages_rendered": len(rendered),
            "by_source": by_source,
            "renderer_counts": renderer_counts,
        },
    }


def _resolve_source_root(
    documents: Sequence[Path],
    source_root: Path | str | None,
) -> Path | None:
    if source_root is not None:
        return Path(source_root).expanduser().resolve()
    if not documents:
        return None
    if len(documents) == 1:
        return documents[0].parent
    common_path = os.path.commonpath([str(path.parent) for path in documents])
    return Path(common_path).resolve()


def _relative_source_name(source: Path, source_root: Path | None) -> str:
    if source_root is not None:
        try:
            return source.relative_to(source_root).as_posix()
        except ValueError:
            pass
    return source.name


def _stable_doc_id(source: Path, source_root: Path | None) -> str:
    relative = _relative_source_name(source.with_suffix(""), source_root)
    slug = re.sub(r"[^a-z0-9]+", "-", relative.lower()).strip("-") or source.stem.lower()
    slug = slug[:48].rstrip("-") or "document"
    digest = hashlib.sha1(str(source).encode("utf-8")).hexdigest()[:10]
    return f"{slug}-{digest}"


def _source_uri(source: Path) -> str:
    return source.resolve().as_uri()


def _build_page_bundle(
    *,
    doc_id: str,
    source: Path,
    relative_source: str,
    pages: Sequence[dict[str, Any]],
) -> dict[str, Any]:
    bundle_pages = []
    for page in pages:
        page_number = int(page["page_number"])
        metadata = dict(page.get("metadata") or {})
        metadata.setdefault("renderer", page["renderer"])
        metadata.setdefault("source_name", source.name)
        metadata.setdefault("source_suffix", source.suffix.lower())
        bundle_pages.append(
            {
                "page_id": f"{doc_id}-page-{page_number}",
                "page_number": page_number,
                "image_path": str(page["image_path"]),
                "source_uri": _source_uri(source),
                "text": page.get("text") or "",
                "width": page.get("width"),
                "height": page.get("height"),
                "metadata": metadata,
            }
        )

    return {
        "bundle_version": DEFAULT_PAGE_BUNDLE_VERSION,
        "doc_id": doc_id,
        "source_uri": _source_uri(source),
        "metadata": {
            "source_path": str(source),
            "relative_source_path": relative_source,
            "page_count": len(bundle_pages),
        },
        "pages": bundle_pages,
    }


def _flatten_bundle_pages(bundle: dict[str, Any]) -> list[dict[str, Any]]:
    flattened = []
    for page in bundle["pages"]:
        flattened.append(
            {
                "source": bundle["metadata"]["source_path"],
                "source_path": bundle["metadata"]["source_path"],
                "source_uri": bundle["source_uri"],
                "image": page["image_path"],
                "image_path": page["image_path"],
                "page": page["page_number"],
                "page_number": page["page_number"],
                "page_id": page["page_id"],
                "doc_id": bundle["doc_id"],
                "bundle_version": bundle["bundle_version"],
                "renderer": page["metadata"].get("renderer", "unknown"),
                "text": page.get("text") or "",
                "width": page.get("width"),
                "height": page.get("height"),
                "metadata": dict(page.get("metadata") or {}),
            }
        )
    return flattened


def _render_source_pages(source: Path, destination_dir: Path) -> list[dict[str, Any]]:
    suffix = source.suffix.lower()
    if suffix in IMAGE_SOURCE_SUFFIXES:
        destination = destination_dir / source.name
        shutil.copy2(source, destination)
        return [
            {
                "page_number": 1,
                "image_path": destination,
                "renderer": "copy",
                "text": source.stem.replace("_", " "),
                "metadata": {},
            }
        ]
    if suffix == ".pdf":
        _require_module("fitz", "PyMuPDF")
        return _render_pdf_pages(source, destination_dir)
    if suffix == ".docx":
        _require_module("docx", "python-docx")
        _require_module("PIL", "Pillow")
        return _render_docx_pages(source, destination_dir)
    if suffix == ".xlsx":
        _require_module("openpyxl", "openpyxl")
        _require_module("PIL", "Pillow")
        return _render_xlsx_pages(source, destination_dir)
    raise ValueError(f"unsupported_suffix:{suffix}")


def _require_module(import_name: str, distribution_name: str) -> None:
    if importlib.util.find_spec(import_name) is None:
        raise MissingDependencyError(distribution_name)


def _render_pdf_pages(source: Path, destination_dir: Path) -> list[dict[str, Any]]:
    import fitz  # type: ignore

    pages = []
    document = fitz.open(source)
    try:
        for page_index in range(document.page_count):
            page = document.load_page(page_index)
            pixmap = page.get_pixmap()
            destination = destination_dir / f"{source.stem}-page-{page_index + 1}.png"
            pixmap.save(destination)
            pages.append(
                {
                    "page_number": page_index + 1,
                    "image_path": destination,
                    "renderer": "pymupdf",
                    "text": page.get_text("text").strip(),
                    "width": pixmap.width,
                    "height": pixmap.height,
                    "metadata": {},
                }
            )
    finally:
        document.close()
    return pages


def _render_docx_pages(source: Path, destination_dir: Path) -> list[dict[str, Any]]:
    from docx import Document  # type: ignore

    document = Document(str(source))
    lines = [paragraph.text.strip() for paragraph in document.paragraphs if paragraph.text.strip()]
    if not lines:
        lines = ["(empty document)"]
    return _render_text_pages(
        lines,
        destination_dir=destination_dir,
        basename=source.stem,
        renderer="python-docx",
    )


def _render_xlsx_pages(source: Path, destination_dir: Path) -> list[dict[str, Any]]:
    import openpyxl  # type: ignore

    workbook = openpyxl.load_workbook(source, data_only=True)
    pages: list[dict[str, Any]] = []
    page_number = 1
    for sheet_index, sheet_name in enumerate(workbook.sheetnames, start=1):
        sheet = workbook[sheet_name]
        rows = []
        for row in sheet.iter_rows(values_only=True):
            values = ["" if cell is None else str(cell) for cell in row]
            if any(value.strip() for value in values):
                rows.append(" | ".join(values))
        if not rows:
            rows = ["(empty sheet)"]
        sheet_pages = _render_text_pages(
            rows,
            destination_dir=destination_dir,
            basename=f"{source.stem}-sheet-{sheet_index}",
            renderer="openpyxl",
            title=sheet_name,
            starting_page_number=page_number,
            metadata={
                "sheet_name": sheet_name,
                "sheet_index": sheet_index,
            },
        )
        pages.extend(sheet_pages)
        page_number += len(sheet_pages)
    return pages


def _render_text_pages(
    lines: Sequence[str],
    *,
    destination_dir: Path,
    basename: str,
    renderer: str,
    title: str | None = None,
    starting_page_number: int = 1,
    metadata: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    from PIL import Image, ImageDraw, ImageFont

    width = 1600
    height = 2000
    margin_x = 48
    margin_y = 48
    line_height = 26
    wrap_width = 112
    font = ImageFont.load_default()
    usable_lines = max(1, (height - (margin_y * 2)) // line_height)

    normalized_lines: list[str] = []
    if title:
        normalized_lines.extend([title, ""])
    for line in lines:
        text = str(line).replace("\t", " ").strip()
        wrapped = textwrap.wrap(text, width=wrap_width) if text else [""]
        normalized_lines.extend(wrapped or [""])
    if not normalized_lines:
        normalized_lines = ["(empty document)"]

    rendered_pages = []
    for page_index, start in enumerate(range(0, len(normalized_lines), usable_lines), start=0):
        page_lines = normalized_lines[start:start + usable_lines]
        image = Image.new("RGB", (width, height), color="white")
        draw = ImageDraw.Draw(image)
        y = margin_y
        for line in page_lines:
            draw.text((margin_x, y), line, fill="black", font=font)
            y += line_height
        destination = destination_dir / f"{basename}-page-{page_index + 1}.png"
        image.save(destination)
        rendered_pages.append(
            {
                "page_number": starting_page_number + page_index,
                "image_path": destination,
                "renderer": renderer,
                "text": "\n".join(line for line in page_lines if line).strip(),
                "width": width,
                "height": height,
                "metadata": dict(metadata or {}),
            }
        )
    return rendered_pages
